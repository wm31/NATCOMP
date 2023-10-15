import numpy as np
import openpyxl

# Define the Rastrigin function
def rastrigin(x):
    d = len(x)
    return 10 * d + sum(x[i]**2 - 10 * np.cos(2 * np.pi * x[i]) for i in range(d))

# PSO Algorithm
def PSO(num_particles, dimensions, num_iterations, target_error, c1, c2, w, bounds):

    # Particle initialization
    class Particle:
        def __init__(self, dimensions):
            self.position = np.array([np.random.uniform(bounds[i][0], bounds[i][1]) for i in range(dimensions)])
            self.velocity = np.array([np.random.uniform(-0.5, 0.5) for i in range(dimensions)])
            self.best_position = np.copy(self.position)
            self.best_score = rastrigin(self.position)

    particles = [Particle(dimensions) for _ in range(num_particles)]
    gbest_position = np.random.uniform(bounds[0][0], bounds[0][1], dimensions)
    gbest_score = rastrigin(gbest_position)

    for particle in particles:
        if particle.best_score < gbest_score:
            gbest_position = np.copy(particle.best_position)
            gbest_score = particle.best_score

    # Main PSO loop
    for iteration in range(num_iterations):
        for particle in particles:

            # Update velocity
            inertia = w * particle.velocity
            personal_attraction = c1 * np.random.random() * (particle.best_position - particle.position)
            global_attraction = c2 * np.random.random() * (gbest_position - particle.position)

            particle.velocity = inertia + personal_attraction + global_attraction

            # Update position
            particle.position += particle.velocity

            # Ensure we stay within bounds
            for i in range(dimensions):
                particle.position[i] = np.clip(particle.position[i], bounds[i][0], bounds[i][1])

            # Update personal best position
            current_score = rastrigin(particle.position)
            if current_score < particle.best_score:
                particle.best_position = np.copy(particle.position)
                particle.best_score = current_score

            # Update global best position
            if current_score < gbest_score:
                gbest_position = np.copy(particle.position)
                gbest_score = current_score

        # Stopping criterion
        if gbest_score < target_error:
            break

    return gbest_position, gbest_score

# Define the ranges of population and dimension sizes to test
pop_sizes = [10, 20, 30]
dim_sizes = [2, 5, 10]

# Define the other PSO parameters
num_iterations = 100
target_error = 1e-6
c1 = 2.0
c2 = 2.0
w = 0.7
bounds = [(-5.12, 5.12)] * dim_sizes[0]

# Create a new Excel workbook and sheet to store the results
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "PSO Results"

# Write the column headings
ws.cell(row=1, column=1, value="Population Size")
ws.cell(row=1, column=2, value="Dimension Size")
ws.cell(row=1, column=3, value="Best Score")
ws.cell(row=1, column=4, value="Best Position")

# Loop over the population and dimension sizes, running PSO for each combination
row_num = 2
for pop_size in pop_sizes:
    for dim_size in dim_sizes:
        bounds = [(-5.12, 5.12)] * dim_size
        best_pos, best_score = PSO(pop_size, dim_size, num_iterations, target_error, c1, c2, w, bounds)
        ws.cell(row=row_num, column=1, value=pop_size)
        ws.cell(row=row_num, column=2, value=dim_size)
        ws.cell(row=row_num, column=3, value=best_score)
        ws.cell(row=row_num, column=4, value=str(best_pos))
        row_num += 1

# Save the workbook to a file
wb.save("PSO_Results.xlsx")